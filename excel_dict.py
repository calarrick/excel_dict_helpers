import openpyxl
from openpyxl.utils.cell import get_column_letter
import logging
from itertools import islice

logger = logging.getLogger(__name__)


def read_sheet(book, sheet, *, start_row=1, col_names = None, has_headers=True, data_only=True, ro=True):
    """"
    Args are the workbook file name and sheet name (as positional arguments, and strings). Optional
    keyword arguments to specify a starting row (col_names, default is 1), select columns to
    include (dict:col_names, default is all columns present in the starting row, and whether the
    starting row is a header row (bool:has_headers).

    Returns a list of dict, each key corresponding to the column name and value to the column
    value in that row.
    If a dict of column letters (keys) and headers (vals) is passed for the
    col_names parameter, only those columns are collected. Otherwise all columns present in the
    selected starting row will be included.

    Instantiates an openpyxl book and sheet using the provided file and sheet names.
    Loops through rows in the openpyxl 'sheet' object. Reads in Excel data to a Python dict.
    (data_only flag on the call to openpyxl is to read last-saved values from
    formula cells, rather than the text of the Excel formulas, which would be the default)"""
    logger.debug('loading from workbook')
    wb = openpyxl.load_workbook(filename=book, data_only=data_only, read_only=ro)
    logger.debug('processing workbook object')
    sheet = wb[sheet]
    p_views = []
    letters = []
    headings = []
    import lxml

    if ro and has_headers:
        for num, row in enumerate(sheet.rows):
            if num == start_row -1:
                # for cell in row:
                #     headings.append(cell.value)
                headings = [h.value for h in row]
            if num >= start_row:
                p_view = {headings[x]: row[x].value for x in range(sheet.min_column-1, sheet.max_column)}
                p_views.append(p_view)
    elif ro:
        for row in islice(sheet.rows, start_row, sheet.max_row):
            p_view = {get_column_letter(x+1): row[x].value for x in range(sheet.min_column-1, sheet.max_column)}
            p_views.append(p_view)

    else:
        if not col_names:
            for num, col in enumerate(sheet.columns):
                col_lett = get_column_letter(num+1)
                if has_headers:
                    col_contents = sheet[str(col_lett) + str(start_row)].value
                else:
                    col_contents = col_lett
                letters.append(col_lett)
                headings.append(col_contents)
            col_names = dict(zip(letters, headings))

        if has_headers:
            start_row = start_row + 1

        for row in range(start_row, sheet.max_row + 1):
            p_view = {}
            for col_lett,col_contents in col_names.items():
                p_view[col_contents] = sheet[str(col_lett) + str(row)].value
            p_views.append(p_view)
        logger.debug('workbook and sheet loaded')
    return p_views


def rewrite_multi_sheet(sheets, out_file_name, extension='.xlsx', wo=False):
    """Provide dict of 'sheet_name: list of rows' and file name, write multiple sheets to new workbook file"""
    if '.' in out_file_name:
        out_file_name = extension[:extension.find('.')]
    wb = openpyxl.Workbook(write_only=wo)
    # will include a blank sheet for sheets without a 'page view' list

    if wo:
        for sheet in sheets:
            ws = wb.create_sheet(str(sheet))
            page_views = sheets[sheet]
            if len(page_views) > 0:
                ws.append([k for k in page_views[0].keys()])
                for view in page_views:
                    ws.append((str(view[k]) for k in view))
    else:
        for sheet in sheets:
            # logger.debug()(dir(sheets[sheet]))
            ws = wb.create_sheet(str(sheet))
            page_views = sheets[sheet]
            logger.debug(page_views)
            if len(sheets[sheet]) > 0:
                try:
                    first_item = page_views[0]
                    col_names = list(first_item.keys())
                    out_row = 1

                    # Build header row
                    for col_num, col_name in enumerate(col_names):
                        ws[get_column_letter(col_num + 1) + str(out_row)].value = col_name

                    # Process all page views from all sessions
                    out_row = 2

                    for view in page_views:
                        for col_num, col_name in enumerate(col_names):
                            ws[get_column_letter(col_num + 1) + str(out_row)].value = str(view[col_name])
                        out_row += 1
                except IndexError as e:
                    logger.warning(str(e))
        del wb['Sheet']

    counter = 0
    try:

        wb.save(out_file_name + extension)
        logger.debug('written to workbook')
    except PermissionError:
        counter +=1
        wb.save(out_file_name + str(counter) + extension)
        logger.debug('written to workbook, incremented  ' + str(counter) + ' for PermissionError, ' + out_file_name)


def rewrite_page_views(page_views, out_file_name, extension = '.xlsx'):
    """Create and save new workbook with one sheet with the processed page view rows"""
    if '.' in out_file_name:
        out_file_name = extension[:extension.find('.')]
    wb = openpyxl.Workbook()
    ws = wb.active
    logger.debug('writing to workbook')

    # Get an item from which to extract the field names (dict keys)
    # BUT NOTE this will break w/ key errors if items don't all have same number of fields
    # pondering value of adding check for that
    first_item = page_views[0]
    col_names = list(first_item.keys())
    out_row = 1
    # Build header row
    for col_num, col_name in enumerate(col_names):
        ws[get_column_letter(col_num + 1) + str(out_row)].value = col_name

    # Process all page views from all sessions
    out_row = 2

    for view in page_views:
        for col_num, col_name in enumerate(col_names):
            ws[get_column_letter(col_num + 1) + str(out_row)].value = str(view[col_name])
        out_row += 1
    saved = 0
    counter = 0
    while saved == 0:
        try:
            if counter > 0:
                increment = str(counter)
            else:
                increment = ''
            wb.save(out_file_name + increment + extension)
            logger.info('written to workbook ' + out_file_name + increment)
            saved = 1
        except PermissionError:
            counter +=1
            logger.info('incrementing ' + str(counter) + ' for PermissionError ' + out_file_name)